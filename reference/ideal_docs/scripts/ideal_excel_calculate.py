"""
ideal_excel_calculate_sheet.py
Mevcut ideal_exe_analiz.xlsx dosyasına "Calculate Metodlari" sayfasi ekler.
"""
import json, re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = r"D:\Projects\_secfix\ideal_analysis\ideal_exe_analiz_v2.xlsx"
JSON = r"D:\Projects\_secfix\ideal_analysis\calculate_methods.json"

# ─── Stil yardımcıları ────────────────────────────────────────────────────
COL = {
    "dark_blue"  : "1B3A5C",
    "mid_blue"   : "2E5F8A",
    "light_blue" : "D6E8F7",
    "green"      : "217346",
    "light_green": "D9EAD3",
    "orange"     : "E87722",
    "purple"     : "6C3483",
    "teal"       : "117A65",
    "red"        : "C0392B",
    "light_red"  : "FADBD8",
    "gold"       : "F0C040",
    "white"      : "FFFFFF",
    "black"      : "000000",
    "gray"       : "F5F5F5",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, cols, bg):
    for c, (w, val) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = fill(bg)
        cell.font      = font(bold=True, color=COL["white"], size=10)
        cell.alignment = center()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(c)].width = w

def row_write(ws, row, values, bg, bold=False):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = fill(bg)
        cell.font      = font(bold=bold, size=10)
        cell.alignment = left()
        cell.border    = thin_border()

def section_hdr(ws, row, text, bg, span):
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill      = fill(bg)
    cell.font      = font(bold=True, color=COL["white"], size=11)
    cell.alignment = left()
    cell.border    = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=span)

# ─── Veri yükle ─────────────────────────────────────────────────────────────
with open(JSON, encoding="utf-8") as f:
    data = json.load(f)

all_records = data["all"]  # list of {class, method, return, args}

# Kategori kural kümesi (metod adından çıkar)
CATEGORY_MAP = [
    (["Bilanco"],               "Bilanço"),
    (["Tominat","Teminat"],     "Teminat"),
    (["Viyop","Viop","Vop"],    "VIOP"),
    (["Imkb","Bist","Hisse"],   "BIST Hisse"),
    (["Takas","TakasGun"],      "BIST Takas"),
    (["Senti","Sentiment"],     "Duyarlılık"),
    (["Macd","Rsi","Stoch","Cci","Adx","Aroon","Atr","Ema","Macd",
       "Momentum","Roc","Rsi","Trix","Ulti","Vert","Swing","Tke","Tti",
       "Qstick","Qsmt","Qqe","Ravi","Mfi","Force","Klinget","Klinger",
       "Kairi","Fore","Demand","Detrend","Ease","Williams","Polarized",
       "Chande","Chaikin","CommodityChannel","CommoditySelection",
       "AccumulationSwing","AccumulationDistribution","RelativeMomentum",
       "RelativeStrength","RelativeVolatility","Rsquared","IntradayMomentum",
       "MassIndex","Stochastic","StandardDeviation","StandardError",
       "ForecastOscillator","FisherTransform","ElliotWave","AwesomeOscill",
       "Pgc","Linear","Projection","PolarizedFractal","RangeInd"],
                                "Momentum / Osilatör"),
    (["MovingAverage","Ema","Dema","Tema","Toma","Mma","HullMA","Tillson",
       "TimeSeries","WeightedClose","TypicalPrice","DoubleMA"],
                                "Ortalama"),
    (["Bollinger","Keltner","AverageTrueRange","Volatility","Envelope",
       "ProjectionBand","StandardErrorBand","ChaikinVolatility","MassIndex"],
                                "Volatilite"),
    (["Pivot","FibonacciBand","Alligator","Ichimoku","ZigZag","PriceChannel",
       "HighLow","PriceRange","ZigZag"],
                                "Trend / Pattern"),
    (["OnBalance","Volume","Hacim","ChankinMoney","ChaikinMoney","MoneyFlow",
       "KlingerOsc","Demand","NegativeVolume","PositiveVolume","EaseOfMove",
       "Accumulation","PriceVolume","VolumeOscil","VolumeSymbol"],
                                "Hacim"),
    (["Faiz","Tufe","Tufetuin"],  "Makro"),
    (["FxSniper","TrendScore","PGC","Pgc","TomaPuan","TTI","TKE",
       "UserEndeks","UserSymbol","Senti","Sembol","Sistem"],
                                "Özel / Kullanıcı"),
    (["Chart","ColLeft","FontH","ButtonPos","CheckSum","Increment",
       "PriceStep","HighLowChange","ChartIndex","DaysToExpiry",
       "TavanTaban","TeorikPrice","Teminat"],
                                "cxBasic Yardımcı"),
    (["Kasa","StartKasa","IslemK","MaxDD","MaxDd","KZ","ViopKZ",
       "ImkbMaliyet","VipMaliyet","ViopPosition","Kademe","ManuelK",
       "Aktarma"],               "Robot / K/Z"),
    (["AKD","Akd","Akdx","HisseAkd","HisseKurum","HisseKz",
       "KurumHisse","KurumTum","DUBIX"],
                                "BIST Analiz"),
]

def get_category(mname):
    for keywords, cat in CATEGORY_MAP:
        for kw in keywords:
            if kw.lower() in mname.lower():
                return cat
    return "Diğer"

# Sınıf → renk (tutarlı)
CLASS_COLORS = {
    "ideal.cxBasic":   ("1B3A5C", "D6E8F7"),
    "ideal.cxSistem":  ("217346", "D9EAD3"),
    "ideal.cxKasa":    ("6C3483", "EAD9F7"),
    "ideal.cxRobot":   ("E87722", "FDE9D0"),
}
DEFAULT_CLASS_COLOR = ("117A65", "D0EDE8")

# Kategoriye göre arka plan
CAT_BG = {
    "Bilanço":           "FFEEBA",
    "Teminat":           "FADBD8",
    "VIOP":              "EAD9F7",
    "BIST Hisse":        "D6E8F7",
    "BIST Takas":        "D6EAF8",
    "Duyarlılık":        "FEF9E7",
    "Momentum / Osilatör": "FCF3CF",
    "Ortalama":          "D9EAD3",
    "Volatilite":        "FDEBD0",
    "Trend / Pattern":   "D5F5E3",
    "Hacim":             "E8DAEF",
    "Makro":             "E9F7EF",
    "Özel / Kullanıcı":  "FDFEFE",
    "cxBasic Yardımcı":  "EBF5FB",
    "Robot / K/Z":       "FAD7A0",
    "BIST Analiz":       "D4E6F1",
    "Diğer":             "F5F5F5",
}

# ─── Workbook aç ─────────────────────────────────────────────────────────────
wb = load_workbook(XLSX)

# Eski sayfayı kaldır (yeniden oluştur)
if "Calculate Metodlari" in wb.sheetnames:
    del wb["Calculate Metodlari"]

ws = wb.create_sheet("Calculate Metodlari")
ws.sheet_view.showGridLines = False

# Başlık
ws.merge_cells("A1:F1")
cell = ws.cell(1, 1, "  ideal.exe — Tüm Calculate Metodları (194 unique / 253 kayıt)")
cell.fill      = fill(COL["dark_blue"])
cell.font      = Font(bold=True, color=COL["white"], size=13, name="Calibri")
cell.alignment = left()
ws.row_dimensions[1].height = 28

r = 3
hdr(ws, r, [
    (32, "Metod Adı"),
    (28, "Sınıf"),
    (20, "Kategori"),
    (18, "Dönüş Tipi"),
    (35, "Parametreler"),
    (10, "#"),
], COL["dark_blue"])
r += 1

# Sırala: kategori → metod adı
sorted_records = sorted(all_records,
    key=lambda d: (get_category(d["method"]), d["method"].lower()))

prev_cat = ""
row_idx = 0

for d in sorted_records:
    mname  = d["method"]
    cname  = d["class"]
    ret    = d["return"]
    args   = d["args"]
    cat    = get_category(mname)
    args_str = ", ".join(args) if args else "—"
    argc   = len(args)

    # Kategori bölüm başlığı
    if cat != prev_cat:
        section_hdr(ws, r, f"  {cat}", COL["mid_blue"], span=6)
        ws.row_dimensions[r].height = 18
        r += 1
        prev_cat = cat
        row_idx = 0

    bg = CAT_BG.get(cat, COL["white"])
    # Satır arası zebra
    if row_idx % 2 == 1:
        # Hafifçe koyulaştır (beyaz kaldır, belirli renkte bırak)
        pass  # CAT_BG zaten renge göre — zebra için sadece beyazla karıştır
    else:
        bg = COL["white"]

    row_write(ws, r, [mname, cname, cat, ret, args_str, argc], bg=bg)

    # Metod adı koyu
    ws.cell(r, 1).font = Font(bold=True, size=10, color="1B3A5C", name="Calibri")
    # Sınıf adı mavi-italik
    ws.cell(r, 2).font = Font(italic=True, size=9, color="2E5F8A", name="Calibri")
    # Kategori ortala
    ws.cell(r, 3).alignment = center()
    # Dönüş tipi
    ws.cell(r, 4).font = Font(color="117A65", size=9, name="Calibri")
    ws.cell(r, 4).alignment = center()
    # Parametre sayısı
    ws.cell(r, 6).alignment = center()
    ws.cell(r, 6).font = Font(bold=True, size=10, color="555555", name="Calibri")

    r += 1
    row_idx += 1

# Sütun genişlikleri
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 36
ws.column_dimensions["F"].width = 8
ws.freeze_panes = "A4"

# ─── Ayrıca: Özet sayfa — Kategorilere göre sayı ─────────────────────────────
# (mevcut Ozet sayfasını güncelleme, basit istatistik satırı)
ws_ozet = wb["Ozet"]

# Sayfanın sonuna boş satır + kategori özeti ekle
from collections import Counter
cat_counts = Counter(get_category(d["method"]) for d in all_records)

max_row = ws_ozet.max_row + 2

section_cell = ws_ozet.cell(max_row, 1, "  Calculate Metod Kategorileri")
section_cell.fill = fill(COL["mid_blue"])
section_cell.font = Font(bold=True, color=COL["white"], size=11, name="Calibri")
section_cell.border = thin_border()
ws_ozet.merge_cells(start_row=max_row, start_column=1,
                    end_row=max_row, end_column=3)
max_row += 1

hdr(ws_ozet, max_row, [(40,"Kategori"),(12,"Kayıt Sayısı"),(25,"Notlar")], COL["mid_blue"])
max_row += 1

for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
    bg = CAT_BG.get(cat, COL["white"])
    row_write(ws_ozet, max_row, [cat, cnt, ""], bg=bg)
    ws_ozet.cell(max_row, 2).alignment = center()
    max_row += 1

# ─── Kaydet ──────────────────────────────────────────────────────────────────
wb.save(XLSX)
import os
print(f"Guncellendi: {XLSX}")
print(f"Boyut: {os.path.getsize(XLSX)/1024:.0f} KB")
print(f"Sayfa sayisi: {len(wb.sheetnames)}")
for sn in wb.sheetnames:
    print(f"  - {sn}")
