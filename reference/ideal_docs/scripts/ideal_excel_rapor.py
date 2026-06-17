"""
ideal.exe Analiz Sonuclari — Excel Raporu
"""
import re, os
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

OUT_DIR = r"D:\Projects\_secfix\ideal_analysis"
XLSX    = r"D:\Projects\_secfix\ideal_analysis\ideal_exe_analiz.xlsx"

# ─── Renk Paleti ─────────────────────────────────────────────────────────────
COL = {
    "dark_blue"  : "1B3A5C",
    "mid_blue"   : "2E5F8A",
    "light_blue" : "D6E8F7",
    "orange"     : "E87722",
    "green"      : "217346",
    "light_green": "D9EAD3",
    "red"        : "C0392B",
    "light_red"  : "FADBD8",
    "gray"       : "F5F5F5",
    "mid_gray"   : "CCCCCC",
    "gold"       : "F0C040",
    "white"      : "FFFFFF",
    "black"      : "000000",
    "purple"     : "6C3483",
    "teal"       : "117A65",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols, bg=COL["dark_blue"], fg=COL["white"], sz=11):
    """Başlık satırı yaz."""
    for c, (col, val) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill   = fill(bg)
        cell.font   = font(bold=True, color=fg, size=sz)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(c)].width = col

def data_row(ws, row, values, bg=COL["white"], bold=False, sz=10):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = fill(bg)
        cell.font      = font(bold=bold, size=sz)
        cell.alignment = left()
        cell.border    = thin_border()

def section_title(ws, row, col, text, bg=COL["mid_blue"], span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill  = fill(bg)
    cell.font  = font(bold=True, color=COL["white"], size=11)
    cell.alignment = left()
    cell.border = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)

# ─── Workbook ────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 1: ÖZET
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Ozet")
ws.sheet_view.showGridLines = False
ws.row_dimensions[1].height = 35

section_title(ws, 1, 1, "  ideal.exe — Veri Yapısı Analiz Raporu", COL["dark_blue"], span=4)
ws.cell(1, 1).font = Font(bold=True, color=COL["white"], size=14, name="Calibri")

r = 3
ws.cell(r, 1, "Analiz Tarihi:").font = font(bold=True, size=10)
ws.cell(r, 2, "11 Nisan 2026").font = font(size=10)
r += 1
ws.cell(r, 1, "Dosya:").font = font(bold=True)
ws.cell(r, 2, r"D:\iDeal\ideal.exe").font = font()
r += 1
ws.cell(r, 1, "Platform:").font = font(bold=True)
ws.cell(r, 2, ".NET Framework 4.6 · x86").font = font()
r += 1
ws.cell(r, 1, "Versiyon:").font = font(bold=True)
ws.cell(r, 2, "0.10.8.6").font = font()
r += 1
ws.cell(r, 1, "Boyut:").font = font(bold=True)
ws.cell(r, 2, "17.5 MB").font = font()
r += 1
ws.cell(r, 1, "MVID:").font = font(bold=True)
ws.cell(r, 2, "54F36D18-35C8-4D4A-A0D4-BC34219BABC3").font = font()
r += 1
ws.cell(r, 1, "Koruma:").font = font(bold=True)
ws.cell(r, 2, "SuppressIldasmAttribute (hafif) — Bilinen obfuscator yok").font = font()

r += 2
section_title(ws, r, 1, "  Tip İstatistikleri", COL["mid_blue"], span=4)
r += 1
header_row(ws, r, [
    (28, "Kategori"), (18, "Adet"), (40, "Açıklama")
], bg=COL["mid_blue"])
r += 1

stats = [
    ("Toplam TypeDef",       3221, "Tüm class/struct/interface/enum tanımları"),
    ("Anlamlı Sınıf",        2126, "Okunabilir isme sahip sınıflar"),
    ("Obfuscated Sınıf",     1060, "Kısaltılmış/hashlenmiş isimler"),
    ("Enum",                   35, "Sabit değer listeleri"),
    ("Veri Modeli Sınıfı",  1117, "Field/Property içeren sınıflar"),
    ("Finans/Trading Sınıfı", 361, "Emir, pozisyon, portföy vb."),
    ("ASCII String",        111088, "Binary içinde bulunan ASCII stringler"),
    ("Unicode String",      26720, "UTF-16LE stringler"),
    ("Finans String",        3068, "Finans/trading ilgili stringler"),
    ("Sütun Adı / Sembol",    878, "Büyük harf sütun adı benzeri stringler"),
]
alt = False
for s in stats:
    bg = COL["light_blue"] if alt else COL["white"]
    data_row(ws, r, list(s), bg=bg)
    ws.cell(r, 2).alignment = center()
    r += 1
    alt = not alt

r += 1
section_title(ws, r, 1, "  Desteklenen Piyasalar", COL["mid_blue"], span=4)
r += 1
header_row(ws, r, [
    (20, "Piyasa"), (15, "Tür"), (20, "Borsa/Platform"), (25, "Notlar")
], bg=COL["mid_blue"])
r += 1
markets = [
    ("BIST Hisse", "Hisse", "IMKB / SASE / TURIB", "Borsa İstanbul"),
    ("VIOP",       "Vadeli/Opsiyon", "VİP", "VOB sözleşmeleri"),
    ("Fon",        "Yatırım Fonu", "TEFAS", "Alım/satım/iptal"),
    ("Kripto Spot","Kripto", "Binance, Icrypex, Artiox", "TRY, USDT, BTC"),
    ("Kripto Future","Kripto Vadeli", "Binance Future, Icrypex","Marjin + leverage"),
    ("Uluslararası","Hisse/Emtia", "NASDAQ, XETRA, EUREX","Dış piyasalar"),
    ("Emtia",      "Emtia", "COMEX, NYMEX","Altın, gümüş, petrol"),
    ("Döviz",      "FX", "EURUSD, ETHHL","Kur verileri"),
]
alt = False
for m in markets:
    bg = COL["light_green"] if alt else COL["white"]
    data_row(ws, r, list(m), bg=bg)
    r += 1
    alt = not alt

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 40
ws.freeze_panes = "A2"

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 2: ENUM DEĞERLERİ
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Enum Degerleri")
ws2.sheet_view.showGridLines = False

section_title(ws2, 1, 1, "  ideal.exe — Enum Tanımları", COL["dark_blue"], span=4)
ws2.cell(1, 1).font = Font(bold=True, color=COL["white"], size=13, name="Calibri")
ws2.merge_cells("A1:D1")
ws2.row_dimensions[1].height = 28

enum_data = [
    ("ideal.OrderSide", "Emir Yönü", [
        ("Buy", 1), ("Sell", 2), ("BuyMinus", 3), ("SellPlus", 4),
        ("SellShort", 5), ("SellShortExempt", 6), ("Undisclosed", 7),
        ("Cross", 8), ("CrossShort", 9), ("CrossShortExempt", 65),
        ("AsDefined", 66), ("Opposite", 67), ("Subscribe", 68),
        ("Redeem", 69), ("Lend", 70), ("Borrow", 71),
    ]),
    ("ideal.OrderType", "Emir Tipi", [
        ("Market", 1), ("Limit", 2), ("Stop", 3), ("StopLimit", 4),
        ("MarketOnClose", 5), ("WithOrWithout", 6), ("LimitOrBetter", 7),
        ("LimitWithOrWithout", 8), ("OnBasis", 9), ("OnClose", 65),
        ("LimitOnClose", 66), ("Funari", 73), ("MarketIfTouched", 74),
        ("Pegged", 80),
    ]),
    ("ideal.OrderStatus", "Emir Durumu", [
        ("PartiallyFilled", 1), ("Filled", 2), ("DoneForDay", 3),
        ("Canceled", 4), ("Replaced", 5), ("PendingCancel", 6),
        ("Stopped", 7), ("Rejected", 8), ("Suspended", 9),
        ("PendingNew", 65), ("Calculated", 66), ("Expired", 67),
        ("AcceptedForBidding", 68), ("PendingReplace", 69),
    ]),
    ("ideal.TimeInForce", "Geçerlilik", [
        ("GoodTillCancel", 1), ("ImmediateOrCancel", 3), ("FillOrKill", 4),
        ("GoodTillDate", 6), ("AtCrossing", 9), ("GoodTillEndOfSession", 83),
    ]),
    ("ideal.MessageType", "Mesaj Tipi (FIX)", [
        ("NewOrder", 65), ("CancelOrder", 66), ("CancelReplaceOrder", 82),
        ("GetOrder", 67), ("GetOrders", 68), ("OrderReport", 69),
        ("Error", 90), ("Heartbeat", 72), ("Login", 76),
        ("LoginResponse", 75), ("Logout", 77), ("GetAccounts", 71),
        ("GetAccountsResponse", 70),
    ]),
    ("ideal.MessageField", "FIX Alan Numaraları", [
        ("Type", 4), ("BodyLength", 5), ("Checksum", 7), ("Id", 11),
        ("Symbol", 12), ("Quantity", 13), ("OrderType", 14), ("OrderSide", 15),
        ("TimeInForce", 16), ("OrderCapacity", 17), ("Account", 18), ("Price", 19),
        ("OrderStatus", 22), ("FilledQuantity", 23), ("RejectReason", 24),
        ("OrderId", 37), ("OrderUuid", 38), ("Username", 34), ("Password", 48),
        ("SendingTime", 52), ("TransactionTime", 54),
    ]),
    ("ideal.OrderCapacity", "Emir Kapasitesi", [
        ("FundOrder", 70), ("Proprietary", 71), ("Individual", 73),
        ("Principal", 80), ("RisklessPrincipal", 82), ("AgentForOtherMember", 87),
    ]),
    ("IDealOrderType", "Robot/Algo Türü", [
        ("Robot", 1), ("iDealGo", 2), ("TaramaRobot", 3), ("RoboTrade", 4),
        ("OtoTrade", 5), ("GridBot", 6), ("TrendBot", 7), ("YatayBot", 8),
        ("PacalBot", 9), ("Arbitraj", 10), ("TrendAlarm", 11),
        ("EgzotikRobot", 12), ("ExecutionAlgo", 13),
    ]),
    ("ideal.AlgoTypes", "Algo Türleri", [
        ("VWAP", 2), ("HCM", 3), ("POV", 4), ("POV2", 5),
        ("ICEBERG", 6), ("ARBITRAJ", 7),
    ]),
    ("ideal.TradeBotStatus", "TradeBot Durumu", [
        ("New", 1), ("Start", 2), ("TP", 3), ("PSL", 4), ("SL", 5), ("Stop", 6),
    ]),
    ("ideal.TradeBotOrderType", "TradeBot Emir Tipi", [
        ("Piyasa", 1),
    ]),
    ("ideal.ConnectionState", "Bağlantı Durumu", [
        ("Connected", 1), ("Authenticating", 2), ("Authenticated", 3),
        ("Failed", 4), ("Error", 5),
    ]),
    ("ideal.enDrawStyles", "Grafik Stili", [
        ("Candle", 1), ("Line", 2), ("Area", 3), ("HeikinAshi", 4), ("Renko", 5),
    ]),
    ("ideal.enAvrMethods", "Ortalama Yöntemi", [
        ("Exponential", 1), ("Weighted", 2), ("Wilder", 3), ("TimeSeries", 4),
        ("Triangular", 5), ("Variable", 6), ("VolumeAdjusted", 7),
        ("ZeroLag", 8), ("HullMA", 9),
    ]),
    ("ideal.enPriceFields", "Fiyat Alanı", [
        ("High", 1), ("Low", 2), ("Close", 3), ("Average", 4), ("Mid", 5), ("Typical", 6),
    ]),
    ("ideal.AccountType", "Hesap Tipi", [
        ("AccountIsCarriedOnNonCustomerSideOfBooks", 2),
        ("HouseTrader", 3), ("FloorTrader", 4),
        ("AccountIsHouseTraderAndIsCrossMargined", 7),
        ("JointBackOfficeAccount", 8),
    ]),
]

# Renk grupları
group_colors = [
    (COL["dark_blue"],  COL["light_blue"]),
    (COL["green"],      COL["light_green"]),
    (COL["orange"],     "FDE9D0"),
    (COL["purple"],     "EAD9F7"),
    (COL["teal"],       "D0EDE8"),
]

r = 3
header_row(ws2, r, [
    (32, "Enum Adı"), (25, "Kategori"), (28, "Değer Adı"), (12, "Sayısal Değer")
], bg=COL["dark_blue"])
r += 1

for idx, (ename, ecat, evals) in enumerate(enum_data):
    hc, lc = group_colors[idx % len(group_colors)]
    # Enum başlığı satırı
    ws2.merge_cells(start_row=r, start_column=1, end_row=r+len(evals)-1, end_column=1)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r+len(evals)-1, end_column=2)
    cell_name = ws2.cell(r, 1, ename)
    cell_name.fill  = fill(hc)
    cell_name.font  = Font(bold=True, color=COL["white"], size=10, name="Calibri")
    cell_name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_name.border = thin_border()
    cell_cat = ws2.cell(r, 2, ecat)
    cell_cat.fill  = fill(hc)
    cell_cat.font  = Font(bold=True, color=COL["white"], size=10, name="Calibri")
    cell_cat.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_cat.border = thin_border()

    for vi, (vname, vnum) in enumerate(evals):
        row_bg = lc if vi % 2 == 0 else COL["white"]
        cell_v = ws2.cell(r + vi, 3, vname)
        cell_v.fill = fill(row_bg)
        cell_v.font = font(size=10)
        cell_v.alignment = left()
        cell_v.border = thin_border()
        cell_n = ws2.cell(r + vi, 4, vnum)
        cell_n.fill = fill(row_bg)
        cell_n.font = font(bold=True, size=10)
        cell_n.alignment = center()
        cell_n.border = thin_border()
    r += len(evals) + 1  # boşluk bırak

ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 34
ws2.column_dimensions["D"].width = 16
ws2.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 3: EMİR MODELLERİ
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Emir Modelleri")
ws3.sheet_view.showGridLines = False
section_title(ws3, 1, 1, "  ideal.exe — Emir Veri Modelleri", COL["dark_blue"], span=5)
ws3.cell(1,1).font = Font(bold=True, color=COL["white"], size=13, name="Calibri")
ws3.merge_cells("A1:E1")
ws3.row_dimensions[1].height = 28

r = 3
header_row(ws3, r, [
    (22, "Sınıf"), (20, "Piyasa"), (8, "Tür"), (30, "Alan Adı"), (18, "Veri Tipi")
], bg=COL["dark_blue"])
r += 1

order_models = [
    ("ImkbOrderRecord", "BIST Hisse", "field", [
        ("LongAccountName","String"), ("AccountName","String"), ("AccountNo","String"),
        ("OrderNo","String"), ("Symbol","String"), ("BuySell","String"),
        ("Amount","R8"), ("AmountShowing","R8"), ("GAmount","R8"), ("Balance","R8"),
        ("GPrice","R8"), ("Price","R8"), ("Total","R8"), ("GTotal","R8"),
        ("ValorDate","String"), ("Status","String"), ("StatusCode","String"),
        ("Session","String"), ("OrderPermit","String"), ("OrderDate","String"),
        ("OrderUpdateDate","String"), ("OrderEndDate","String"), ("OrderType","String"),
        ("CancelPermit","String"), ("AmendPermit","String"), ("OneSessionPermit","String"),
        ("OrderRef","String"), ("OrderSessionNo","String"), ("ZincirRef","String"),
        ("Note","String"), ("Validity","String"), ("SatisTip","String"),
        ("GSaat","String"), ("EmirUpdateNum","I4"), ("SiraNo","I4"),
        ("MaxZincirSiraNo","I4"), ("RefNo","String"), ("BorsaRefNo","String"),
        ("SessionName","String"), ("ExecutionStatus","String"), ("Selected","UI1"),
        ("[prop] OrderNoString","String"),
    ]),
    ("VipOrderRecord", "VIOP", "field", [
        ("LongAccountName","String"), ("AccountName","String"), ("AccountNo","String"),
        ("OrderNo","String"), ("RecordNo","String"), ("Symbol","String"),
        ("BuySell","String"), ("SubMarket","String"),
        ("Amount","R8"), ("GAmount","R8"), ("Balance","R8"), ("GPrice","R8"),
        ("Price","R8"), ("Stop","R8"), ("Total","R8"), ("GTotal","Decimal"),
        ("ValorDate","String"), ("Status","String"), ("StatusCode","String"),
        ("State","String"), ("CancelReason","String"), ("PositionClosing","String"),
        ("BorsaDurum","String"), ("OrderDate","String"), ("OrderTime","String"),
        ("OrderType","String"), ("EndDate","String"), ("PriceType","String"),
        ("CancelPermit","String"), ("AmendPermit","String"),
        ("BorsaEmirNo","String"), ("TemsilciRef","String"),
        ("EnteredAmount","R8"), ("InvisibleAmount","R8"), ("VisibleBalance","R8"),
        ("OrderRef","String"), ("SartTip","String"), ("SartYon","String"),
        ("SartSembol","String"), ("SartFiyat","R8"), ("SartBool","Boolean"),
        ("SessionName","String"), ("ExecutionStatus","String"), ("Selected","UI1"),
    ]),
    ("CriptoOrderRecord", "Kripto Spot", "field", [
        ("AccountName","String"), ("AccountNo","String"), ("OrderNo","String"),
        ("Symbol","String"), ("Kod","String"), ("BuySell","String"),
        ("Amount","Decimal"), ("GAmount","R8"), ("Balance","R8"),
        ("GPrice","R8"), ("Price","Decimal"), ("StopPrice","Decimal"),
        ("Total","R8"), ("GTotal","R8"), ("ValorDate","String"),
        ("Status","String"), ("StatusCode","String"), ("Session","String"),
        ("OrderType","String"), ("ZincirRef","String"), ("Note","String"),
        ("Validity","String"), ("SatisTip","String"), ("ExecutionStatus","String"),
        ("Selected","UI1"), ("[prop] OrderNoString","String"),
    ]),
    ("NewOrder", "API / FIX", "property", [
        ("OrderUuid","String"), ("Symbol","String"), ("Quantity","R8"),
        ("OrderType","OrderType enum"), ("OrderSide","OrderSide enum"),
        ("TimeInForce","TimeInForce enum"), ("OrderCapacity","Nullable<OrderCapacity>"),
        ("Account","String"), ("AccountType","Nullable<AccountType>"),
        ("Price","Nullable<R8>"), ("OffHoursTrading","Boolean"),
    ]),
    ("CancelOrder", "API / FIX", "property", [
        ("OrderUuid","String"), ("OrderId","I4"), ("Symbol","String"),
        ("Quantity","R8"), ("OrderSide","OrderSide enum"), ("Account","String"),
    ]),
    ("CancelReplaceOrder", "API / FIX", "property", [
        ("OrderUuid","String"), ("OrderId","I4"), ("Symbol","String"),
        ("Quantity","R8"), ("Price","R8"), ("OrderSide","OrderSide enum"),
        ("OrderType","OrderType enum"), ("Account","String"),
    ]),
    ("ReqEquitySendOrder", "BIST API", "property", [
        ("appCode","String"), ("appPassword","String"), ("accountId","String"),
        ("clOrdId","String"), ("instrumentSymbol","String"), ("uniqueSymbol","String"),
        ("instrumentType","String"), ("qty","R8"), ("price","R8"),
        ("sideId","String"), ("closeShortSell","Boolean"), ("orderTypeId","String"),
        ("timeInForceId","String"), ("token","String"), ("tokenVersion","String"),
        ("marketSegmentAlert","String"), ("maxFloor","R8"),
    ]),
    ("VipSendOrderReq", "VIOP API", "property", [
        ("sozlesme","String"), ("islem","String"), ("miktar","I4"),
        ("fiyat","Decimal"), ("orderType","String"), ("sureTarih","String"),
        ("sure","String"), ("gorunenMiktar","Decimal"),
        ("tetikTipi","I4"), ("tetikFiyat","I4"), ("tetikSozlesme","Object"),
        ("acikKapali","I4"), ("smsGonderimi","String"), ("aksamSeansi","String"),
    ]),
    ("OrderReport", "FIX Yanıt", "property", [
        ("OrderId","I4"), ("OrderStatus","OrderStatus enum"),
        ("OriginalOrderId","Nullable<I4>"), ("FilledQuantity","Nullable<R8>"),
        ("ExecutedQuantity","Nullable<R8>"), ("BuyPrice","Nullable<R8>"),
        ("RejectReason","String"), ("TransactionTime","String"),
        ("OrderSendingTime","String"),
    ]),
]

model_colors = [
    (COL["dark_blue"],  COL["light_blue"]),
    (COL["purple"],     "EAD9F7"),
    ("1A5276",          "D6EAF8"),
    (COL["green"],      COL["light_green"]),
    (COL["orange"],     "FDE9D0"),
    (COL["teal"],       "D0EDE8"),
    ("6E2F8B",          "E8DAEF"),
    ("7D6608",          "FCF3CF"),
    (COL["red"],        COL["light_red"]),
]

for mi, (mname, mpiyasa, mkind, mfields) in enumerate(order_models):
    hc, lc = model_colors[mi % len(model_colors)]
    ws3.merge_cells(start_row=r, start_column=1, end_row=r+len(mfields)-1, end_column=1)
    ws3.merge_cells(start_row=r, start_column=2, end_row=r+len(mfields)-1, end_column=2)
    ws3.merge_cells(start_row=r, start_column=3, end_row=r+len(mfields)-1, end_column=3)
    for col_idx, val in [(1, mname), (2, mpiyasa), (3, mkind.upper())]:
        c = ws3.cell(r, col_idx, val)
        c.fill = fill(hc)
        c.font = Font(bold=True, color=COL["white"], size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    for fi, (fname, ftype) in enumerate(mfields):
        row_bg = lc if fi % 2 == 0 else COL["white"]
        cf = ws3.cell(r+fi, 4, fname)
        cf.fill = fill(row_bg); cf.font = font(size=10)
        cf.alignment = left(); cf.border = thin_border()
        ct = ws3.cell(r+fi, 5, ftype)
        ct.fill = fill(row_bg);  ct.font = font(color="1A5276", size=10)
        ct.alignment = center(); ct.border = thin_border()
    r += len(mfields) + 1

ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 32
ws3.column_dimensions["E"].width = 22
ws3.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 4: POZİSYON MODELLERİ
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Pozisyon Modelleri")
ws4.sheet_view.showGridLines = False
section_title(ws4, 1, 1, "  ideal.exe — Pozisyon & Portföy Veri Modelleri", COL["dark_blue"], span=5)
ws4.cell(1,1).font = Font(bold=True, color=COL["white"], size=13, name="Calibri")
ws4.merge_cells("A1:E1")
ws4.row_dimensions[1].height = 28

r = 3
header_row(ws4, r, [
    (24, "Sınıf"), (18, "Piyasa"), (30, "Alan Adı"), (16, "Veri Tipi"), (35, "Açıklama")
], bg=COL["dark_blue"])
r += 1

pos_models = [
    ("ImkbPositionRecord", "BIST Hisse", [
        ("Symbol","String","Hisse kodu"),
        ("Lot","R8","Toplam lot"),
        ("LastPrice","R8","Son fiyat"),
        ("Sellable","R8","Satılabilir lot"),
        ("Cost","R8","Alış maliyeti"),
        ("Bloke","R8","Bloke miktar"),
        ("ProfitX","R8","Kâr/zarar"),
        ("PortfolioType","String","Portföy tipi"),
        ("uniqueSymbol","String","Benzersiz sembol"),
        ("equityType","String","Hisse tipi"),
        ("balanceT / T1 / T2 / T3","R8","Valörlü bakiyeler"),
        ("avgPrice","R8","Ortalama fiyat"),
        ("depotCode","String","Depo kodu"),
        ("currentAmount","R8","Güncel tutar"),
        ("AssetType","String","Varlık tipi"),
        ("BalanceType","String","Bakiye tipi"),
        ("LotT1 / LotT2","R8","Valörlü lotlar"),
        ("PortfoyOran","R8","Portföy oranı (%)"),
        ("Rumuz","String","Hesap rumuz"),
        ("DovizCinsi","String","Döviz cinsi"),
        ("DovizDegeri","R8","Döviz değeri"),
        ("[prop] Profit","R8","TL kâr/zarar"),
        ("[prop] ProfitYuzde","R8","Yüzde kâr/zarar"),
        ("[prop] TotalTL","R8","Toplam TL değer"),
    ]),
    ("VipPositionRecord", "VIOP", [
        ("Symbol","String","Sözleşme kodu"),
        ("BuyAmount / SellAmount","R8","Alış/satış adedi"),
        ("OpenAmount","R8","Açık pozisyon adedi"),
        ("NetAmount","R8","Net pozisyon"),
        ("UnitAmount","R8","Birim adedi"),
        ("OpenPosition","R8","Açık pozisyon"),
        ("Profit","R8","Kâr/zarar"),
        ("PozSize","Decimal","Pozisyon büyüklüğü"),
        ("ProfitAnlik","Decimal","Anlık kâr"),
        ("ProfitFifo","Decimal","FIFO kâr"),
        ("SonUzlasi","R8","Son uzlaşı fiyatı"),
        ("Status","String","Durum"),
        ("Direction","String","Yön (Long/Short)"),
        ("Price","R8","Fiyat"),
        ("LastPrice","R8","Son fiyat"),
        ("SettlementPrice","R8","Uzlaşı fiyatı"),
        ("ContractType","String","Sözleşme tipi"),
        ("Tip","String","Tip"),
        ("Risk","String","Risk durumu"),
        ("Currency","String","Para birimi"),
        ("Cost","R8","Maliyet"),
        ("NetFifoMaliyet","R8","Net FIFO maliyet"),
        ("OpsiyonPrimiNet","R8","Net opsiyon primi"),
        ("FifoMaliyet / AcilisMaliyet","R8","FIFO / açılış maliyet"),
        ("balanceT/T1/T2/T3","R8","Valörlü bakiyeler"),
        ("avgPrice","R8","Ortalama fiyat"),
        ("qty","R4","Miktar"),
    ]),
    ("CriptoPositionRecord", "Kripto Spot", [
        ("Symbol","String","Sembol"),
        ("Coin","String","Coin adı"),
        ("Description","String","Açıklama"),
        ("Lot","R8","Adet"),
        ("Locked","R8","Kilitli miktar"),
        ("LastPrice","R8","Son fiyat"),
        ("Sellable","R8","Satılabilir"),
        ("Cost","R8","Maliyet"),
        ("Bloke","R8","Bloke"),
        ("ProfitX","R8","Kâr/zarar"),
        ("Available","R8","Kullanılabilir"),
        ("AssetType","String","Varlık tipi"),
        ("BalanceType","String","Bakiye tipi"),
        ("[prop] Total","R8","Toplam değer"),
        ("[prop] Profit","R8","Kâr/zarar"),
    ]),
    ("FonPositionRecord", "Yatırım Fonu", [
        ("FonAdi","String","Fon adı"),
        ("FonKodu","String","Fon kodu"),
        ("Adet","R4","Adet"),
        ("SatilabilirAdet","R4","Satılabilir adet"),
        ("Maliyet","R4","Alış maliyeti"),
        ("DegerlendirmeFiyati","R4","Değerleme fiyatı"),
        ("VarlikTutari","R4","Varlık tutarı"),
        ("KarZarar","R4","Kâr/zarar"),
        ("PortfoyOrani","R8","Portföy oranı"),
    ]),
    ("ViopPozisyon", "VIOP (API)", [
        ("sozlesmeKodu","String","Sözleşme kodu"),
        ("uzunKisa","String","L/S yön"),
        ("pozisyonSayisi","I4","Pozisyon sayısı"),
        ("maliyet","R4","Maliyet"),
        ("kapanis","R4","Kapanış fiyatı"),
        ("parasalTutar","R4","Parasal tutar"),
        ("guniciKz","R4","Gün içi K/Z"),
    ]),
    ("Bar", "Grafik/Tarihsel", [
        ("time","R4","Zaman damgası"),
        ("open","R4","Açılış"),
        ("close","R4","Kapanış"),
        ("high","R4","Yüksek"),
        ("low","R4","Düşük"),
        ("volume","R4","Hacim"),
    ]),
    ("Getchartdata", "API Chart", [
        ("ts","I8","Unix timestamp"),
        ("open","R8","Açılış"),
        ("high","R8","Yüksek"),
        ("low","R8","Düşük"),
        ("close","R8","Kapanış"),
        ("volume","R8","Hacim"),
    ]),
]

for pi, (pname, ppiyasa, pfields) in enumerate(pos_models):
    hc, lc = model_colors[pi % len(model_colors)]
    ws4.merge_cells(start_row=r, start_column=1, end_row=r+len(pfields)-1, end_column=1)
    ws4.merge_cells(start_row=r, start_column=2, end_row=r+len(pfields)-1, end_column=2)
    for col_idx, val in [(1, pname), (2, ppiyasa)]:
        c = ws4.cell(r, col_idx, val)
        c.fill = fill(hc)
        c.font = Font(bold=True, color=COL["white"], size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    for fi, (fname, ftype, fdesc) in enumerate(pfields):
        row_bg = lc if fi % 2 == 0 else COL["white"]
        ws4.cell(r+fi, 3, fname).fill = fill(row_bg)
        ws4.cell(r+fi, 3).font = font(size=10)
        ws4.cell(r+fi, 3).alignment = left()
        ws4.cell(r+fi, 3).border = thin_border()
        ws4.cell(r+fi, 4, ftype).fill = fill(row_bg)
        ws4.cell(r+fi, 4).font = font(color="1A5276", size=10)
        ws4.cell(r+fi, 4).alignment = center()
        ws4.cell(r+fi, 4).border = thin_border()
        ws4.cell(r+fi, 5, fdesc).fill = fill(row_bg)
        ws4.cell(r+fi, 5).font = font(color="555555", size=10)
        ws4.cell(r+fi, 5).alignment = left()
        ws4.cell(r+fi, 5).border = thin_border()
    r += len(pfields) + 1

ws4.column_dimensions["A"].width = 26
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 38
ws4.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 5: ANA PORTFÖY SINIFI
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Portfoy Sinifi")
ws5.sheet_view.showGridLines = False
section_title(ws5, 1, 1, "  class Portfoy — Ana Veri Ambarı", COL["dark_blue"], span=4)
ws5.cell(1,1).font = Font(bold=True, color=COL["white"], size=13, name="Calibri")
ws5.merge_cells("A1:D1")
ws5.row_dimensions[1].height = 28

r = 3
header_row(ws5, r, [
    (38, "Alan Adı"), (28, "Veri Tipi (kısaltılmış)"), (16, "Grup"), (35, "Açıklama")
], bg=COL["dark_blue"])
r += 1

portfoy_fields = [
    # BIST
    ("ImkbPositionList",        "List<ImkbPositionRecord>",    "BIST", "BIST açık pozisyonlar"),
    ("ImkbOrderList",           "List<ImkbOrderRecord>",        "BIST", "BIST bekleyen emirler"),
    ("ImkbStatementList",       "List<ImkbStatementRecord>",    "BIST", "BIST ekstre"),
    ("ImkbTransactionReports",  "List<ImkbTransactionReport>",  "BIST", "BIST işlem raporları"),
    ("ImkbSummaryDictionary",   "Dict<String,String>",          "BIST", "Özet bilgiler"),
    ("ImkbStockSellableDictionary","Dict<String,R8>",           "BIST", "Satılabilir lot dict"),
    ("ImkbStockLimitDictionary","Dict<String,R8>",              "BIST", "Limit dict"),
    ("ImkbLimit",               "R8",                           "BIST", "İşlem limiti"),
    ("ImkbOverall",             "R8",                           "BIST", "Overall bakiye"),
    ("ImkbCariBakiye",          "R8",                           "BIST", "Cari bakiye"),
    ("ImkbKrediDahilLimit",     "R8",                           "BIST", "Kredi dahil limit"),
    ("ImkbKrediBorcu",          "R8",                           "BIST", "Kredi borcu"),
    ("ImkbBakiyeFarkNet",       "R8",                           "BIST", "Bakiye fark (net)"),
    ("ImkbBakiyeFarkYuzde",     "R8",                           "BIST", "Bakiye fark (%)"),
    # VIOP
    ("VipPositionList",         "List<VipPositionRecord>",      "VIOP", "VİOP açık pozisyonlar"),
    ("VipOrderList",            "List<VipOrderRecord>",          "VIOP", "VİOP bekleyen emirler"),
    ("VipGerceklesenList",      "List<VipOrderRecord>",          "VIOP", "VİOP gerçekleşen"),
    ("ViopTeminatToplam",       "R8",                            "VIOP", "Teminat toplam"),
    ("ViopTeminatBaslangic",    "R8",                            "VIOP", "Başlangıç teminatı"),
    ("ViopTeminatSurdurme",     "R8",                            "VIOP", "Sürdürme teminatı"),
    ("ViopTeminatKullanilabilir","R8",                           "VIOP", "Kullanılabilir teminat"),
    ("ViopTeminatCekilebilir",  "R8",                            "VIOP", "Çekilebilir teminat"),
    ("ViopNetMaliyet",          "R8",                            "VIOP", "Net maliyet"),
    ("ViopProfitLoss",          "R8",                            "VIOP", "Opsiyonlar kâr/zarar"),
    ("ToplamTeminat",           "R8",                            "VIOP", "Toplam teminat"),
    ("GayriNakdiTeminat",       "R8",                            "VIOP", "Gayri nakdi teminat"),
    # Kripto
    ("CriptoOrderList",         "List<CriptoOrderRecord>",      "Kripto", "Kripto emirler"),
    ("CriptoPositionList",      "List<CriptoPositionRecord>",   "Kripto", "Kripto spot pozisyon"),
    ("CriptoTradeList",         "List<CriptoTradeRecord>",      "Kripto", "Kripto geçmiş işlem"),
    ("CriptoFuturePositions",   "List<Position>",               "Kripto", "Kripto vadeli pozisyon"),
    ("CriptoFutureOpenOrders",  "List<OpenOrder>",              "Kripto", "Kripto vadeli bekleyen em."),
    ("IcrypexFutureAssetList",  "List<Assets>",                 "Kripto", "Icrypex vadeli varlıklar"),
    # Fon
    ("FonPositionList",         "List<FonPositionRecord>",      "Fon", "Fon pozisyonları"),
    ("FonIslemList",            "List<FonIslemRecord>",         "Fon", "Fon işlemleri"),
    ("FonTanimList",            "List<FonTanimRecord>",         "Fon", "Fon tanımları"),
    ("FonKurucuDict",           "Dict<String,String>",          "Fon", "Kurucu dict"),
    # Diğer
    ("VarlikList",              "List<VarlikRecord>",           "Diğer", "Varlık listesi"),
    ("DovizBakiye",             "R8",                           "Diğer", "Döviz bakiye"),
    ("MaksimumPortfoyDegerLimit","R8",                          "Diğer", "Maks portföy değer limit"),
    ("KullanilanPortfoyDegeri", "R8",                           "Diğer", "Kullanılan portföy değeri"),
    ("StopOutRiskOrani",        "R8",                           "Diğer", "Stop-out risk oranı"),
]

group_bg = {
    "BIST":   (COL["light_blue"], COL["dark_blue"]),
    "VIOP":   (COL["light_green"], COL["green"]),
    "Kripto": ("FDE9D0", COL["orange"]),
    "Fon":    ("EAD9F7", COL["purple"]),
    "Diğer":  ("D0EDE8", COL["teal"]),
}
prev_group = ""
for fi, (fname, ftype, fgroup, fdesc) in enumerate(portfoy_fields):
    if fgroup != prev_group:
        section_title(ws5, r, 1, f"  {fgroup}", group_bg[fgroup][1], span=4)
        ws5.row_dimensions[r].height = 20
        r += 1
        prev_group = fgroup
    lc, _ = group_bg[fgroup]
    row_bg = lc if fi % 2 == 0 else COL["white"]
    data_row(ws5, r, [fname, ftype, fgroup, fdesc], bg=row_bg)
    ws5.cell(r, 2).font = font(color="1A5276", size=10)
    ws5.cell(r, 3).alignment = center()
    r += 1

ws5.column_dimensions["A"].width = 40
ws5.column_dimensions["B"].width = 30
ws5.column_dimensions["C"].width = 12
ws5.column_dimensions["D"].width = 40
ws5.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 6: FİNANSAL RASYOLAR & HESAP MODELLERİ
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Rasyo ve Hesap")
ws6.sheet_view.showGridLines = False
section_title(ws6, 1, 1, "  Finansal Rasyolar & Hesap Modelleri", COL["dark_blue"], span=4)
ws6.cell(1,1).font = Font(bold=True, color=COL["white"], size=13, name="Calibri")
ws6.merge_cells("A1:D1")
ws6.row_dimensions[1].height = 28

r = 3
section_title(ws6, r, 1, "  HisseRasyoRecord — Bilanço & Finansal Rasyolar", COL["green"], span=4)
r += 1
header_row(ws6, r, [
    (30, "Alan Adı"), (12, "Tip"), (38, "Açıklama"), (20, "Kategori")
], bg=COL["green"])
r += 1

rasyo_fields = [
    ("Donem","String","Dönem (e.g. 2025/Q4)","Kimlik"),
    ("Symbol","String","Hisse kodu","Kimlik"),
    ("sector","String","Sektör","Kimlik"),
    ("CariOran","R8","Cari oran (Dönen Var / Kısa Vadeli Borç)","Likidite"),
    ("LikitOran","R8","Likit oran","Likidite"),
    ("NakitOran","R8","Nakit oran","Likidite"),
    ("KaldiracOran","R8","Kaldıraç oranı","Borç"),
    ("BorcOzsermayeOran","R8","Borç/Özsermaye","Borç"),
    ("MadDurVarOzserOran","R8","Maddi duran varlık/özsermaye","Borç"),
    ("KVadeBorTopBorOran","R8","Kısa vadeli borç / Toplam borç","Borç"),
    ("AlacakDevHiz","R8","Alacak devir hızı","Faaliyet"),
    ("AlacakTahsilSuresi","R8","Alacak tahsil süresi (gün)","Faaliyet"),
    ("StokDevirHiz","R8","Stok devir hızı","Faaliyet"),
    ("StokKalmaSure","R8","Stok kalma süresi (gün)","Faaliyet"),
    ("NetIsSerDevHiz","R8","Net işletme sermayesi devir hızı","Faaliyet"),
    ("OzkaynakDevHiz","R8","Özkaynak devir hızı","Faaliyet"),
    ("AktifDevHiz","R8","Aktif devir hızı","Faaliyet"),
    ("TicBorcDevHiz","R8","Ticari borç devir hızı","Faaliyet"),
    ("TicBorcOdeSure","R8","Ticari borç ödeme süresi","Faaliyet"),
    ("BurutKarMarj","R8","Brüt kâr marjı (%)","Kârlılık"),
    ("NetKarMarj","R8","Net kâr marjı (%)","Kârlılık"),
    ("FaliyetKarMarj","R8","Faaliyet kâr marjı (%)","Kârlılık"),
    ("AktifKarlilik","R8","Aktif kârlılık (ROA)","Kârlılık"),
    ("OzSerKarlilik","R8","Özsermaye kârlılığı (ROE)","Kârlılık"),
    ("EsasFaaliyetKari","R8","Esas faaliyet kârı","Kârlılık"),
    ("FAVOK","R8","FAVÖK (EBITDA)","Kârlılık"),
    ("NetDonemKar","R8","Net dönem kârı","Kârlılık"),
    ("NetIsletmeSer","R8","Net işletme sermayesi","Kârlılık"),
]

for fi, row in enumerate(rasyo_fields):
    bg = COL["light_green"] if fi % 2 == 0 else COL["white"]
    data_row(ws6, r, list(row), bg=bg)
    ws6.cell(r, 2).alignment = center()
    ws6.cell(r, 4).alignment = center()
    r += 1

r += 1
section_title(ws6, r, 1, "  HesapListesiData — Hesap Bilgileri", COL["mid_blue"], span=4)
r += 1
header_row(ws6, r, [(30,"Alan"),(12,"Tip"),(38,"Açıklama"),(20,"")], bg=COL["mid_blue"])
r += 1
hesap_fields = [
    ("ADI","String","Müşteri adı",""),
    ("SOYADI","String","Müşteri soyadı",""),
    ("VirmanliSatisButonu","I4","Virmanlı satış butonu",""),
    ("AltPazarRBF","I4","Alt pazar RBF",""),
    ("OTOMATIK_ALIM_SATIM_SZL","I4","Otomatik alım satım sözleşmesi",""),
    ("VERI_DAGITIM_OZEL_EKRAN","I4","Veri dağıtım özel ekran",""),
    ("ReturnValue","I4","Dönüş kodu",""),
    ("errorMessage","String","Hata mesajı",""),
    ("errorUniqueCode","I4","Benzersiz hata kodu",""),
]
for fi, row in enumerate(hesap_fields):
    bg = COL["light_blue"] if fi % 2 == 0 else COL["white"]
    data_row(ws6, r, list(row), bg=bg)
    ws6.cell(r, 2).alignment = center()
    r += 1

r += 1
section_title(ws6, r, 1, "  ViopRobotHesapClass & BistRobotHesapClass — Robot Hesabı", COL["orange"], span=4)
r += 1
header_row(ws6, r, [(30,"Alan"),(12,"Tip"),(38,"Açıklama"),(20,"Sınıf")], bg=COL["orange"])
r += 1
robot_fields = [
    ("Pozisyonlar","List<ImkbPositionRecord>","Açık BIST pozisyonlar","BistRobotHesapClass"),
    ("GerceklesenEmirler","List<ImkbOrderRecord>","Gerçekleşen BIST emirler","BistRobotHesapClass"),
    ("BekleyenEmirler","List<ImkbOrderRecord>","Bekleyen BIST emirler","BistRobotHesapClass"),
    ("IslemLimit","R8","İşlem limiti (TL)","BistRobotHesapClass"),
    ("Bakiye","R8","Nakit bakiye","BistRobotHesapClass"),
    ("Pozisyonlar","List<VipPositionRecord>","Açık VIOP pozisyonlar","ViopRobotHesapClass"),
    ("GerceklesenEmirler","List<VipOrderRecord>","Gerçekleşen VIOP emirler","ViopRobotHesapClass"),
    ("BekleyenEmirler","List<VipOrderRecord>","Bekleyen VIOP emirler","ViopRobotHesapClass"),
    ("TeminatToplam","R8","Toplam teminat","ViopRobotHesapClass"),
    ("TeminatBaslangic","R8","Başlangıç teminatı","ViopRobotHesapClass"),
    ("TeminatSurdurme","R8","Sürdürme teminatı","ViopRobotHesapClass"),
    ("TeminatKullanilabilir","R8","Kullanılabilir teminat","ViopRobotHesapClass"),
    ("TeminatCekilebilir","R8","Çekilebilir teminat","ViopRobotHesapClass"),
    ("TeminatCagri","R8","Teminat çağrısı","ViopRobotHesapClass"),
]
for fi, row in enumerate(robot_fields):
    bg = "FDE9D0" if fi % 2 == 0 else COL["white"]
    data_row(ws6, r, list(row), bg=bg)
    ws6.cell(r, 2).font = font(color="1A5276", size=10)
    ws6.cell(r, 2).alignment = center()
    ws6.cell(r, 4).alignment = center()
    r += 1

ws6.column_dimensions["A"].width = 34
ws6.column_dimensions["B"].width = 26
ws6.column_dimensions["C"].width = 42
ws6.column_dimensions["D"].width = 24
ws6.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 7: İNDİKATÖR LİSTESİ
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Indiktorler (137)")
ws7.sheet_view.showGridLines = False
section_title(ws7, 1, 1, "  ideal.exIndicatorTypes — 137 Teknik İndikatör", COL["dark_blue"], span=4)
ws7.cell(1,1).font = Font(bold=True, color=COL["white"], size=13, name="Calibri")
ws7.merge_cells("A1:D1")
ws7.row_dimensions[1].height = 28

r = 3
header_row(ws7, r, [
    (8, "No"), (32, "İndikatör Adı"), (28, "Kategori"), (40, "Açıklama")
], bg=COL["dark_blue"])
r += 1

indicators = [
    (1,"Alligator","Trend","Williams Alligator"),
    (2,"BollingerBands","Volatilite","Bollinger Bantları"),
    (3,"Dema","Ortalama","Çift üstel hareketli ortalama"),
    (4,"Envelope","Ortalama","Zarf"),
    (5,"FibonacciBand","Fibonacci","Fibonacci bantları"),
    (6,"HighLowRange","Fiyat","Yüksek-Düşük aralığı"),
    (7,"Ichimoku","Trend","Ichimoku Kinko Hyo"),
    (8,"KeltnerChannel","Volatilite","Keltner kanalı"),
    (9,"LinearRegression","İstatistik","Doğrusal regresyon"),
    (10,"MovingAverage","Ortalama","Hareketli ortalama"),
    (11,"ParabolicSAR","Trend","Parabolik SAR"),
    (12,"PivotBand","Destek/Direnç","Pivot bantları"),
    (13,"PriceChannel","Trend","Fiyat kanalı"),
    (14,"ProjectionBands","Volatilite","Projeksiyon bantları"),
    (15,"StandardErrorBands","İstatistik","Standart hata bantları"),
    (16,"Tema","Ortalama","Üçlü üstel hareketli ortalama"),
    (17,"TimeSeriesForecast","İstatistik","Zaman serisi tahmini"),
    (18,"Toma","Ortalama","TOMA"),
    (19,"TypicalPrice","Fiyat","Tipik fiyat"),
    (20,"WeightedClose","Fiyat","Ağırlıklı kapanış"),
    (21,"ZigZagPercent","Trend","Zigzag (yüzde)"),
    (22,"ZigZagPoints","Trend","Zigzag (puan)"),
    (23,"AccumulationDistribution","Hacim","Birikim/Dağıtım"),
    (24,"AccumulationSwingIndex","Hacim","Birikim swing endeksi"),
    (25,"AroonUpDown","Momentum","Aroon Yukarı/Aşağı"),
    (26,"AroonOscillator","Momentum","Aroon osilatörü"),
    (27,"AverageDirectionalIndex","Trend","ADX"),
    (28,"AverageDirectionalRating","Trend","ADR"),
    (29,"AverageTrueRange","Volatilite","ATR"),
    (30,"AwesomeOscillator","Momentum","Harika Osilatör"),
    (31,"BilancoFK","Bilanço","Faaliyete Konulanlar"),
    (32,"BilancoNetKar","Bilanço","Net Kâr"),
    (33,"BilancoOdenmisSerm","Bilanço","Ödenmiş Sermaye"),
    (34,"BilancoOzSerm","Bilanço","Özsermaye"),
    (35,"BilancoPD","Bilanço","Piyasa Değeri"),
    (36,"BilancoPDDD","Bilanço","PD/DD"),
    (37,"BollingerWidth","Volatilite","Bollinger genişliği"),
    (38,"ChaikinMoneyFlow","Hacim","Chaikin Para Akışı"),
    (39,"ChaikinOscillator","Hacim","Chaikin Osilatörü"),
    (40,"ChaikinVolatility","Volatilite","Chaikin Volatilite"),
    (41,"ChandeMomentum","Momentum","Chande Momentum"),
    (42,"CommodityChannelIndex","Momentum","CCI"),
    (43,"CommoditySelectionIndex","Momentum","CSI"),
    (44,"DemandIndex","Hacim","Talep Endeksi"),
    (45,"DetrendedPriceOscillator","Momentum","Detrended fiyat osc."),
    (46,"DirectionalIndicator","Trend","Yönsel gösterge"),
    (47,"DirectionalMovement","Trend","Yönsel hareket"),
    (48,"EaseOfMovement","Hacim","Hareket kolaylığı"),
    (49,"ForecastOscillator","İstatistik","Tahmin osilatörü"),
    (50,"IntradayMomentumIndex","Momentum","Güniçi momentum"),
    (51,"Kairi","Momentum","Kairi relatif endeksi"),
    (52,"KlingerOscillator","Hacim","Klinger osilatörü"),
    (53,"LinearRegressionIndicator","İstatistik","Doğrusal regresyon göst."),
    (54,"LinearRegressionSlope","İstatistik","Doğrusal regresyon eğimi"),
    (55,"Lot","Hacim","Lot"),
    (56,"Macd","Momentum","MACD"),
    (57,"MacdHistogram","Momentum","MACD histogram"),
    (58,"MassIndex","Volatilite","Kütle endeksi"),
    (59,"Momentum","Momentum","Momentum"),
    (60,"MoneyFlowIndex","Hacim","Para akış endeksi (MFI)"),
    (61,"NegativeVolumeIndex","Hacim","Negatif hacim endeksi"),
    (62,"OnBalanceVolume","Hacim","OBV"),
    (63,"OpenInterest","VIOP","Açık pozisyon"),
    (64,"Performance","Performans","Performans"),
    (65,"PolarizedFractalEfficiency","Trend","Polarize fraktal eff."),
    (66,"PositiveVolumeIndex","Hacim","Pozitif hacim endeksi"),
    (67,"PriceOscillatorPercent","Momentum","Fiyat osc. (yüzde)"),
    (68,"PriceOscillatorPercentHistogram","Momentum","Fiyat osc. hist. (yüzde)"),
    (69,"PriceOscillatorPoints","Momentum","Fiyat osc. (puan)"),
    (70,"PriceOscillatorPointsHistogram","Momentum","Fiyat osc. hist. (puan)"),
    (71,"PriceRocPercent","Momentum","Fiyat ROC (yüzde)"),
    (72,"PriceRocPoints","Momentum","Fiyat ROC (puan)"),
    (73,"PriceVolumeTrend","Hacim","Fiyat-hacim trendi"),
    (74,"ProjectionBandwidth","Volatilite","Projeksiyon bant genişliği"),
    (75,"ProjectionOscillator","Volatilite","Projeksiyon osilatörü"),
    (76,"Qstick","Momentum","Qstick"),
    (77,"RangeIndicator","Volatilite","Aralık göstergesi"),
    (78,"RelativeMomentumIndex","Momentum","RMI"),
    (79,"RelativeStrengthIndex","Momentum","RSI"),
    (80,"RelativeVolatilityIndex","Volatilite","RVI"),
    (81,"Rsquared","İstatistik","R-kare"),
    (82,"StandardDeviation","İstatistik","Standart sapma"),
    (83,"StandardError","İstatistik","Standart hata"),
    (84,"StochasticMomentumIndex","Momentum","Stochastic momentum"),
    (85,"StochasticOscillator","Momentum","Stochastic"),
    (86,"StochasticFastOscillator","Momentum","Stochastic hızlı"),
    (87,"StochasticRSI","Momentum","Stochastic RSI"),
    (88,"SwingIndex","Momentum","Swing endeksi"),
    (89,"Takas","BIST","Takas verisi"),
    (90,"TKE","Momentum","TKE"),
    (91,"TrendScore","Trend","Trend skoru"),
    (92,"Trix","Momentum","TRIX"),
    (93,"UltimateOscillator","Momentum","Ultimate osilatör"),
    (94,"VerticalHorizontalFilter","Trend","VHF"),
    (95,"Volume","Hacim","Hacim"),
    (96,"VolumeOscillatorPercent","Hacim","Hacim osc. (yüzde)"),
    (97,"VolumeOscillatorPercentHistogram","Hacim","Hacim osc. hist. (yüzde)"),
    (98,"VolumeOscillatorPoints","Hacim","Hacim osc. (puan)"),
    (99,"VolumeOscillatorPointsHistogram","Hacim","Hacim osc. hist. (puan)"),
    (100,"VolumeSymbol","Hacim","Sembol hacmi"),
    (101,"VolumeSymbolPercent","Hacim","Sembol hacim (yüzde)"),
    (102,"WilliamsAccDist","Hacim","Williams birikim/dağıtım"),
    (103,"WilliamsR","Momentum","Williams %R"),
    (104,"HighLowBox","Fiyat","Yüksek-Düşük kutusu"),
    (105,"DoubleMA","Ortalama","Çift hareketli ortalama"),
    (106,"PHPL01","BIST","PH/PL01"),
    (107,"EhlersFilter","Filtre","Ehlers filtresi"),
    (108,"EhlersDistCoefFilter","Filtre","Ehlers dist. katsayı filtresi"),
    (109,"RAVI","Momentum","RAVI"),
    (110,"RSIDenvelope","Momentum","RSI zarf"),
    (111,"FxSniper","Sinyal","FxSniper"),
    (112,"TomaPuan","TOMA","TOMA puanı"),
    (113,"QuantitativeQualitativeEstimation","Momentum","QQE"),
    (114,"DigerSymbol","Fiyat","Diğer sembol fiyatı"),
    (115,"HHV","Fiyat","Dönemsel yüksek"),
    (116,"LLV","Fiyat","Dönemsel düşük"),
    (117,"ElliotWaveOscillator","Momentum","Elliot dalga osilatörü"),
    (118,"StochasticSlow","Momentum","Stochastic yavaş"),
    (119,"TTI","Momentum","TTI"),
    (120,"MMA","Ortalama","MMA"),
    (121,"KurumAnaliz","BIST","Kurum analizi"),
    (122,"TillsonT3","Ortalama","Tillson T3"),
    (123,"HullMA","Ortalama","Hull MA"),
    (124,"TakasGun","BIST","Takas günü"),
    (125,"TakasDegisim","BIST","Takas değişim"),
    (126,"Sentiment","Duygu","Piyasa duyarllılığı"),
    (127,"Senti60","Duygu","60 dk. duyarlılık"),
    (128,"SentiGun","Duygu","Günlük duyarlılık"),
    (129,"SentiMom60","Duygu","60 dk. duyarlılık mom."),
    (130,"SentiMomGun","Duygu","Günlük duyarlılık mom."),
    (131,"SentiOscilator","Duygu","Duyarlılık osilatörü"),
    (132,"Pivot","Destek/Direnç","Pivot noktaları"),
    (133,"Senti15","Duygu","15 dk. duyarlılık"),
    (134,"PGC","Momentum","PGC"),
    (135,"FaizMb","Makro","Merkez Bankası faizi"),
    (136,"TufeTuik","Makro","TÜFE (TÜİK)"),
    (137,"FisherTransform","Momentum","Fisher dönüşümü"),
]

cat_colors = {
    "Trend":"D6EAF8", "Ortalama":"D9EAD3", "Volatilite":"FADBD8",
    "Momentum":"FCF3CF", "Hacim":"E8DAEF", "Bilanço":"FFEEBA",
    "İstatistik":"D5F5E3", "Fiyat":"E8F8F5", "BIST":"D6EAF8",
    "VIOP":"E8DAEF", "Filtre":"F9EBEA", "Sinyal":"FBEEE6",
    "TOMA":"E8F8F5", "Duygu":"FEF9E7", "Makro":"E9F7EF",
    "Destek/Direnç":"F4ECF7", "Fibonacci":"FDF2F8", "Performans": "EBF5FB",
}

for no, iname, icat, idesc in indicators:
    bg = cat_colors.get(icat, COL["white"])
    data_row(ws7, r, [no, iname, icat, idesc], bg=bg)
    ws7.cell(r, 1).alignment = center()
    ws7.cell(r, 3).alignment = center()
    r += 1

ws7.column_dimensions["A"].width = 8
ws7.column_dimensions["B"].width = 36
ws7.column_dimensions["C"].width = 20
ws7.column_dimensions["D"].width = 42
ws7.freeze_panes = "A4"

# ─── Kaydet ──────────────────────────────────────────────────────────────────
wb.save(XLSX)
print(f"Excel dosyasi olusturuldu:")
print(f"  {XLSX}")
import os
print(f"  Boyut: {os.path.getsize(XLSX)/1024:.0f} KB")
